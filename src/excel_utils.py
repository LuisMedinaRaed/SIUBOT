import openpyxl
import os

from juicio_utils import Juicio

def leer_datos_excel(archivo, hoja):
    """
    Lee valores desde un archivo Excel.

    Args:
    - archivo (str): Ruta del archivo Excel.
    - hoja (str): Nombre de la hoja en la que se deben leer los datos.

    Returns:
    - list: Lista de listas que contiene los valores de la hoja especificada.
    """
    datos = []
    try:
        # Abre el archivo Excel
        libro_excel = openpyxl.load_workbook(archivo, data_only=True)  # Utiliza data_only=True para obtener valores, no fórmulas
        # Selecciona la hoja especificada
        hoja_excel = libro_excel[hoja]

        # Lee los valores de cada celda en la hoja
        for fila in hoja_excel.iter_rows(values_only=True):
            # Comprueba si algún valor en la fila es None
            if None not in fila:
                datos.append(list(fila))

        return datos
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        return []

def formatear_datos_a_juicios(datos_leidos):
    """
    Formatea los datos leídos en una lista de objetos Juicio.

    Parámetros:
    - datos_leidos (list): Una lista de listas que contiene datos, donde cada lista representa una fila.

    Retorna:
    - list: Una lista de objetos Juicio, donde cada objeto Juicio se crea a partir de una fila en los datos leídos.
    """
    # Lista para almacenar objetos Juicio
    juicios = []

    # Iterar sobre las filas del array
    for fila in datos_leidos[1:]:  # Excluye la primera fila que contiene los nombres de las columnas
        # Crear un objeto Juicio y agregarlo a la lista
        juicio = Juicio(*fila)
        juicios.append(juicio)

    return juicios

def es_archivo_excel(ruta_archivo):
    """
    Verifica si la ruta del archivo es un archivo de Excel válido.

    Parámetros:
    - ruta_archivo (str): La ruta del archivo a verificar.

    Retorna:
    - bool: True si es un archivo de Excel válido, False en caso contrario.
    """
    if not ruta_archivo or not os.path.exists(ruta_archivo):
        return False

    try:
        workbook = openpyxl.load_workbook(ruta_archivo)
        return True
    except Exception as e:
        return False